from tkinter import *
from tkinter import messagebox     
import datetime
from openpyxl import Workbook,load_workbook
import pandas as pd
vaqt=datetime.datetime.now()

top = Tk()  
top.geometry("800x600")
#______________________________________
wb=Workbook()
ws=wb.active
w1=load_workbook('Baza.xlsx')
wk=w1.active
 
    
#----------maxsulot qo'shish-----------
def clear1():
    nom1.delete(0,END)
    son1.delete(0,END)
    narx1.delete(0,END)
    kod1.delete(0,END)
    yil1.delete(0,END)
    oy1.delete(0,END)
    kun1.delete(0,END)

def saqlash():
            wb=Workbook()
            ws=wb.active
            w1=load_workbook('Baza.xlsx')
            wk=w1.active
            data=[]
            for i in range(1, wk.max_row + 1):
                    A = f"A{i}"
                    B = f"B{i}"
                    C = f"C{i}"
                    D = f"D{i}"
                    E = f"D{i}"
                    F = f"F{i}"
                    G = f"G{i}"
                    data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value,wk[G].value])
    
           
            if nom1.get()==0:
                messagebox.showinfo("Eslatma","ma'lumotlarni to'liq kiriting")
            else:
                nomi=nom1.get()
            if  son1.get()==0:
                messagebox.showinfo("Eslatma","ma'lumotlarni to'liq kiriting")
            else:
                soni=son1.get()
            
            narx=narx1.get()
                
            foiz=int(narx)*1.2    # stafka 20%
            code_sh=kod1.get()
            now=datetime.datetime.now().strftime('%d/%m/%y')

            yil=(yil1.get())
               
            oy=(oy1.get())
                
            kun=(kun1.get())
            if nomi==False :
                        
                  messagebox.showinfo("Eslatma","ma'lumotlarni to'liq kiriting")
            else:        
                muddat=datetime.datetime(int(yil),int(oy),int(kun)).strftime('%d/%m/%y')
                   
                for k in data:
                    d=data.index(k)
                    ws.delete_rows(d+1)
                        
                          
                data.append([nomi,soni,narx,foiz,code_sh,now,muddat])
                for k in data:
                    ws.append(k)
                wb.save("Baza.xlsx")
                messagebox.showinfo("Massage","maxsulot muofaqiyatli qo'shildi  ✔️")

                clear1()
                    
            
            


def qoshish():
    global nom1,son1,narx1,kod1,yil1,oy1,kun1
    global frame
    frame = Frame(top,bg='gray85',width=1500,height=600)  
    frame.place(x=0,y=0)
    
    nom= Label(top, text = "Nomi",font=('Times New Roman', 17),bg='gray85',height='1',width='7')
    nom.place(x = 40,y = 90)
    
    son= Label(top, text = "Soni ",font=('Times New Roman', 16),bg='gray85',height='1',width='7')
    son.place(x = 40,y = 135)
    
    narx= Label(top, text = "Narxi",font=('Times New Roman', 16),bg='gray85',height='1',width='7')
    narx.place(x = 40,y = 180)
        
    kod= Label(top, text = "Shtrix kodi",font=('Times New Roman', 16),bg='gray85',height='1',width='8')
    kod.place(x = 30,y = 225)
    
    yil= Label(top, text = "Yil",font=('Times New Roman', 16),bg='gray85',height='1',width='7')
    yil.place(x = 40,y = 270)
    
    oy= Label(top, text = "Oy",font=('Times New Roman', 16),bg='gray85',height='1',width='7')
    oy.place(x = 40,y = 315)
    
    kun= Label(top, text = "Kun",font=('Times New Roman', 16),bg='gray85',height='1',width='7')
    kun.place(x = 40,y = 360)


    nom1= Entry(top,width = 28,font=('Helvetica', 15))
    nom1.place(x = 150, y = 90)
    
    son1 = Entry(top,width = 28,font=('Helvetica', 15))
    son1.place(x = 150, y = 135)
    
    narx1 = Entry(top,width = 28,font=('Helvetica', 15))
    narx1.place(x = 150, y = 180)
    
    kod1 = Entry(top,width = 28,font=('Helvetica', 15))
    kod1.place(x = 150, y = 225)
    
    yil1 = Entry(top,width = 28,font=('Helvetica', 15))
    yil1.place(x = 150, y = 270)
    
    oy1 = Entry(top,width = 28,font=('Helvetica', 15))
    oy1.place(x = 150, y = 315)
    
    kun1 = Entry(top,width = 28,font=('Helvetica', 15))
    kun1.place(x = 150, y = 360)
    
    but1 = Button(top, text = " CLEAR ",bg='lemonchiffon',activebackground = "wheat",fg='black',activeforeground = "blue",font=('Times New Roman', 17, 'bold'),command=clear1)
    but1.place(x = 460, y = 440)
    
    but2 = Button(top, text = "  OK   ",bg='lemonchiffon',activebackground = "wheat",fg='black', activeforeground = "blue",font=('Times New Roman', 17, 'bold'),command=saqlash)
    but2.place(x =600, y = 440)

    lab2= Label(top, text = "Yangi maxsulotlarni qo'shish",font=('Times New Roman', 16),bg='gray85',height='1',width='60')
    lab2.place(x = 20,y = 40)
    
    #========= ko'rish funksiyasi=============
def korish():
    global frame
    frame = Frame(top,bg='gray85',width=800,height=600)  
    frame.place(x=0,y=0)
    #scrollbar = Scrollbar(top,bd=500,width=20,elementborderwidth=1)
    #scrollbar.place(x=760,y=10 )#side = RIGHT, fill = Y )
    #scrollbar.config( command = Text.yview )
    
    df = pd.read_excel("Baza.xlsx")
      
    n_rows = df.shape[0]
    n_cols = df.shape[1]
    
    column_names = df.columns
    i=0
    for j, col in enumerate(column_names):
        text = Text(top, width=13, height=1, bg = "#9BC2E6")
        text.grid(row=i,column=j)
        text.insert(INSERT, col)
          
    
    for i in range(n_rows):
        for j in range(n_cols):
            text = Text(top, width=13, height=1)#,yscrollcommand = scrollbar.set)
            text.grid(row=i+1,column=j)
            text.insert(INSERT, df.loc[i][j])

        
    
    #==========sotish funksiyasi===========

               
   
def clear3():
    
    nomi3.delete(0,END)
    soni3.delete(0,END)
    
def  sot():
    wb=Workbook()
    ws=wb.active
    w1=load_workbook('Baza.xlsx')
    wk=w1.active
    data=[]
    kx=[]
    sotish=[]
    for i in range(1, wk.max_row + 1):
                    A = f"A{i}"
                    B = f"B{i}"
                    C = f"C{i}"
                    D = f"D{i}"
                    E = f"D{i}"
                    F = f"F{i}"
                    G = f"G{i}"
                    data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value,wk[G].value])

               
    ismi=nomi3.get()
    df=0
    for k in data:
        if str(ismi) == k[0]:
            df=0
            son=soni3.get()
            
            if int(son)<=int(k[1]):
                k[1]=int(k[1])-int(son)
                soni2 = Label(top,text=f"{k[0]} dan {k[1]} ta bor ",bg='gray99',font=('Times New Roman', 16))
                soni2.place(x = 150, y = 155)
                for h in data:
                    ws.append(h)                                   
                wb.save("Baza.xlsx")
                break                
            elif int(son) > int(k[1]):
                messagebox.showinfo("Eslatma","maxsulot soni kam!!!")
            break
        else:
            df+=1
    if df==len(data):
        messagebox.showinfo("Eslatma","bunday maxsulot yo'q!!!")


    
    
def sotish():
    global frame
    frame = Frame(top,bg='gray85',width=800,height=600)  
    frame.place(x=0,y=0)
    
    lab2= Label(top, text = "Maxsulotlarni sotish",font=('Times New Roman', 16),bg='gray85',height='1',width='50')
    lab2.place(x = 20,y = 30)
    lab3= Label(top, text = "Maxsulot nomini kiriting",font=('Times New Roman', 16),bg='gray85',height='1',width='50')
    lab3.place(x = 20,y = 60)
    # label
    sotish_nomi= Label(top, text = "Nomi",font=('Times New Roman', 17),bg='gray85',height='1',width='7')
    sotish_nomi.place(x = 40,y = 110)
    
    sotish_soni= Label(top, text = "Soni ",font=('Times New Roman', 16),bg='gray85',height='1',width='7')
    sotish_soni.place(x = 40,y = 200)
     
    
    # entry
    global nomi3,soni3
    nomi3= Entry(top,width =25,font=('Helvetica', 15))
    nomi3.place(x = 150, y = 110)

    soni3= Entry(top,width = 8,font=('Helvetica', 15))
    soni3.place(x = 150, y = 200)

    sotish_but1 = Button(top, text = " CLEAR ",bg='lemonchiffon',activebackground = "wheat",fg='black',activeforeground = "blue",font=('Times New Roman', 17, 'bold'),command=clear3)
    sotish_but1.place(x = 150, y = 440)
    
    sotish_but2 = Button(top, text = "  OK   ",bg='lemonchiffon',activebackground = "wheat",fg='black', activeforeground = "blue",font=('Times New Roman', 17, 'bold'),command=sot)
    sotish_but2.place(x =550, y = 440)

    #======================








    
def muddat():
    global frame
    frame.destroy()
    frame = Frame(top,bg='gray85',width=800,height=600)  
    frame.place(x=0,y=0)
    #scrollbar = Scrollbar(top)
    #scrollbar.pack( side = RIGHT, fill = Y )
    #scrollbar.config( command = Text.yview )
   
    df = pd.read_excel("Baza.xlsx")
      
    n_rows = df.shape[0]
    n_cols = df.shape[1]
    
    column_names = df.columns
    i=0
    for j, col in enumerate(column_names):
        text = Text(top, width=13, height=1, bg = "#9BC2E6")
        text.grid(row=i,column=j)
        text.insert(INSERT, col)
          
    
    for i in range(n_rows):
        for j in range(n_cols):
            
            
            y=(df.loc[i][6].split('/'))
            ww=datetime.datetime(int('20'+y[2]),int(y[1]),int(y[0]))-datetime.datetime.now()
            srok=ww.days
            
            if srok<=0:                
                text = Text(top, width=13, height=1)#,yscrollcommand = scrollbar.set )
                text.grid(row=i+1,column=j)
                text.insert(INSERT, df.loc[i][j])
                continue
    if srok>0:
        messagebox.showinfo("Eslatma","Muddati o'tgan maxsulotlar topilmadi!!!")

 




    
        





    
    

    #======================
    
def xisobot():
    global frame
    frame = Frame(top,bg='gray85',width=800,height=600)  
    frame.place(x=0,y=0)
    b1 = Button(top,text = "xisobot",width = 8,font=('Times New Roman', 16, 'bold'),fg='blue',bg='lemonchiffon',command= korish)
    b1.place(x=100,y=100)

    #======================
    
def yordam():
    global frame
    frame = Frame(top,bg='gray85',width=800,height=600)  
    frame.place(x=0,y=0)
    lab= Label(top, text = "bu dastur exe fayli bilan ishlaydi",font=('Times New Roman', 16),bg='lemonchiffon',height='1',width='60')
    lab.place(x = 10,y = 250)

    
    #-------------------------- 
sotish()




menubar = Menu(top) 
menubar.add_command(label="➕ Maxsulot qo'shish", command=qoshish)  
menubar.add_command(label="👁 Maxsulotlarni ko'rish", command=korish)
menubar.add_command(label="🛒 Maxulot sotish",command= sotish)  
menubar.add_command(label="⏳ Muddati o'tganlar", command=muddat) 
menubar.add_command(label="💼 Xisobot", command=xisobot)
menubar.add_command(label="📒 Yordam", command=yordam)  
 
top.config(menu=menubar)

top.mainloop()  



           
     
    
    

   




